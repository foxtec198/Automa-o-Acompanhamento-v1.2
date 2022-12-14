from openpyxl import load_workbook as lw
from sqlite3 import connect
from tkinter import *
from tkinter import messagebox, PhotoImage
from time import strftime as st

from metas import Metas
from vj import VJ
from pa import PA
from hxh import mainHxh

from functools import partial
import os
import matplotlib.pyplot as plt
# TESTE DO GIT HUB

class App():

    def __init__(self):
        # Vars
        self.win = Tk()
        self.conn = connect("dados.db")
        self.c = self.conn.cursor()
        self.caulcarGrafico()
        self.bg = "#333"
        self.nome = 'Undefined'
        self.dia = int(st("%d"))
        m = Metas()
        self.metaLoja = m.soma

        # callback
        self.config_win()
        self.widget()
        self.inst()
        self.win.mainloop()

    def config_win(self):
        self.win.geometry("1300x680+0+0")
        self.win['bg'] = 'Black'
        self.win.iconbitmap("img/icon.ico")
        self.win.title("Automação Acompanhamento")
        self.win.resizable(width=False, height=False)

    def widget(self):  # All Widgets
        win = self.win
        # Frames
        self.loginFrame = Frame(win, bg=self.bg, width=210, height=660)

        self.hxhFrame = Frame(win, bg=self.bg, width=260, height=80)
        self.paFrame = Frame(win, bg=self.bg, width=260, height=80)
        self.vjFrame = Frame(win, bg=self.bg, width=260, height=80)
        self.esFrame = Frame(win, bg=self.bg, width=250, height=80)

        self.txtFrame = Frame(win, bg=self.bg, width=550, height=540)
        self.graphFrame = Frame(win, bg=self.bg, width=500, height=540)
        # x = width
        # y = height

        pimg = PhotoImage
        # IMAGENS
        self.imglogin = pimg(file=r"img\\man.png")
        self.paImg = pimg(file=r"img\\pa.png")
        self.hxhImg = pimg(file=r"img\\hxh.png")

        self.vjImg = pimg(file=r"img\\vj.png")
        self.userImg = pimg(file=r"img\\user.png")
        self.setImg = pimg(file=r"img\\set.png")
        self.cartImg = pimg(file=r"img\\cart.png")
        self.g1 = pimg(file=r"img\\meta.png")
        self.g2 = pimg(file=r"img\\real.png")

        # DIREITOS AUTORAIS ----------------------------------------------------------------------------------------
        self.ms = 'Deselvolvido por TecnoBreve Enterprise, Direitos autorais reservados a mesma © Londrina 2022'
        Label(self.win,
              text=self.ms,
              font='Arial 8 italic',
              bg='black',
              fg='grey').place(x=690, y=660, anchor='center')

        # Login
        # Vermelho: d90429
        # Back: #333
        self.lblImgLogin = Label(self.loginFrame,
                                 image=self.imglogin,
                                 bg=self.bg,
                                 width=150,
                                 height=150)
        self.lblUser = Label(self.loginFrame,
                             text='👤 Undefined',
                             bg=self.bg,
                             fg='white',
                             font='sansserif 12 bold')
        self.lblHora = Label(self.loginFrame,
                             font="sansserif 12",
                             bg=self.bg,
                             fg="white")
        self.atualizar()
        self.lblMeta = Label(self.loginFrame,
                             text=f"Meta Loja: R$ {self.metaLoja:.2f}",
                             font="Arial 12 bold",
                             bg=self.bg,
                             fg='white')
        self.btnLogin = Button(self.loginFrame,
                               text="Login",
                               width=25,
                               font="monospace 16 bold",
                               bg="Black",
                               fg="ghostwhite",
                               command=self.login)
        self.btnLogout = Button(self.loginFrame,
                                text="Logout",
                                width=25,
                                font="monospace 16",
                                bg="#d90429",
                                fg="white",
                                command=self.logout)

        # TEXT FRAME
        self.txt = Text(self.txtFrame,
                        bg=self.bg,
                        fg='white',
                        borderwidth=3,
                        height=30,
                        width=65)
        Label(self.txtFrame,
              text='🖊 Cole o texto aqui:',
              font='Arial 14 bold',
              bg=self.bg,
              fg="GhostWhite").place(x=280, y=15, anchor='center')
        self.btnClear = Button(self.txtFrame,
                               text='Limpar',
                               bg='#d90429',
                               fg='white',
                               width=5,
                               height=1,
                               activebackground='#333',
                               command=self.cls1)

        # HxH
        self.btnHxh = Button(self.hxhFrame,
                             text='HoraXhora',
                             image=self.hxhImg,
                             width='64',
                             bg=self.bg,
                             borderwidth=0,
                             activebackground=self.bg,
                             command=self.hxh)
        self.lblHxh = Label(self.hxhFrame,
                            text='Parcial \nHora X Hora',
                            font='monospace 14 bold',
                            bg=self.bg,
                            fg='white')

        self.btnPA = Button(self.paFrame,
                            text='Participalçao',
                            image=self.paImg,
                            width='64',
                            bg=self.bg,
                            borderwidth=0,
                            activebackground=self.bg,
                            command=self.pa)
        self.lblPA = Label(self.paFrame,
                           text='Parcial \nParticipação',
                           font='monospace 14 bold',
                           bg=self.bg,
                           fg='white')

        self.btnVJ = Button(self.vjFrame,
                            text='VendaComJuros',
                            image=self.vjImg,
                            width='64',
                            bg=self.bg,
                            borderwidth=0,
                            activebackground=self.bg,
                            command=self.vj)
        self.lblVJ = Label(self.vjFrame,
                           text='Parcial \nVendas com Juros',
                           font='monospace 14 bold',
                           bg=self.bg,
                           fg='white')

        self.btnCart = Button(self.esFrame,
                              text='E-Store',
                              image=self.cartImg,
                              width='64',
                              bg=self.bg,
                              borderwidth=0,
                              activebackground=self.bg,
                              command=self.estore)
        self.lblEs = Label(self.esFrame,
                           text='E-store',
                           font='monospace 14 bold',
                           bg=self.bg,
                           fg='white')

        self.btnAdd = Button(self.loginFrame,
                             text='AdicionarUser',
                             image=self.userImg,
                             width='64',
                             bg=self.bg,
                             borderwidth=0,
                             activebackground=self.bg,
                             command=self.addUser)
        self.btnSet = Button(self.loginFrame,
                             text='Settings',
                             image=self.setImg,
                             width='64',
                             bg=self.bg,
                             borderwidth=0,
                             activebackground=self.bg,
                             command=self.settings)

        # GRAPH FRAME
        fgg = self.graphFrame
        self.gp1 = Label(fgg, image=self.g1, bg=self.bg)
        self.gp2 = Label(fgg, image=self.g2, bg=self.bg)

    def inst(self):  # Instanciar
        # Frame
        self.loginFrame.place(x=10, y=10)

        self.hxhFrame.place(x=230, y=10)
        self.paFrame.place(x=500, y=10)
        self.vjFrame.place(x=770, y=10)
        self.esFrame.place(x=1040, y=10)
        self.txtFrame.place(x=230, y=101)

        self.graphFrame.place(x=790, y=101)

        # Login
        self.lblImgLogin.place(x=25, y=10)
        self.lblUser.place(x=100, y=160, anchor="center")
        self.lblHora.place(x=100, y=180, anchor="center")
        self.lblMeta.place(x=100, y=220, anchor="center")
        self.btnLogin.place(x=100, y=600, anchor="center")
        self.btnLogout.place(x=100, y=640, anchor="center")

        # Text
        self.txt.place(x=10, y=30)
        self.btnClear.place(x=490, y=2)

        # DP
        self.btnHxh.place(x=10, y=10)
        self.lblHxh.place(x=170, y=40, anchor='center')
        self.btnPA.place(x=10, y=10)
        self.lblPA.place(x=170, y=40, anchor='center')
        self.btnVJ.place(x=10, y=10)
        self.lblVJ.place(x=170, y=40, anchor='center')
        self.btnCart.place(x=10, y=10)
        self.lblEs.place(x=170, y=40, anchor='center')

        #self.btnAdd.place(x = 40, y = 500)
        #self.btnSet.place(x = 110, y = 502)

        # FG
        self.gp1.place(x=0, y=10)
        self.gp2.place(x=0, y=300)

    # Functions
    def msg(self, tp, msg):
        if tp == 1:
            messagebox.showinfo('Auto Acompanhamento', msg)
        elif tp == 2:
            messagebox.showerror('Auto Acompanhamento', msg)
        elif tp == 3:
            messagebox.showwarning('Auto Acompanhamento', msg)

    # CALLBACKS FUNCTIONS EXT ============================+
    def hxh(self):
        txt = self.txt.get('1.0', END)
        arquivo = open('texto.txt', 'w')
        arquivo.write(txt)
        arquivo.close()

        self.msg(1, 'Abrindo parcial')
        mainHxh()
        os.system('horaxhora.xlsx')
        self.atug()

    def pa(self):
        txt = self.txt.get('1.0', END)
        arquivo = open('clb.txt', 'w')
        arquivo.write(txt)
        arquivo.close()

        PA().cons()
        self.msg(1, 'Parcial realizado com sucesso!')
        os.system('parcial.xlsx')

    def vj(self):
        txt = self.txt.get('1.0', END)
        arquivo = open('vj.txt', 'w')
        arquivo.write(txt)
        arquivo.close()

        VJ().code()
        self.msg(1, 'Abrido parcial')
        os.system('vj.xlsx')

    def estore(self):
        if self.nome != 'Undefined':
            try:
                self.msg(1, 'Abrindo estore')
                os.system('estore.xlsx')
                os.system('estore.py')
            except:
                self.msg(2, 'Erro, procure um desenvolvedor')
        else:
            self.msg(2, 'Usuário não permitido, Realize login!')

    #+====================================================+
    def atug(self):
        self.caulcarGrafico()
        self.g2 = PhotoImage(file=r"img\\real.png")
        self.gp2 = Label(self.graphFrame, image=self.g2, bg=self.bg)
        self.gp2.place(x=0, y=300)

    def atualizar(self):
        self.hora = st("%H:%M:%S - %d/%m/%Y")
        self.lblHora["text"] = self.hora
        self.lblHora.after(100, self.atualizar)

    def cls(self, event):
        self.txt.delete("1.0", END)

    def cls1(self):
        self.txt.delete("1.0", END)

    def login(self):
        win2 = Tk()
        win2.iconbitmap("img/icon.ico")
        win2.geometry("250x250")
        win2['bg'] = '#333'
        win2.resizable(width=False, height=False)
        win2.title("LOGIN")

        def cons(event):
            self.userg = self.user.get()
            self.senhag = self.senha.get()
            user = self.c.execute(
                f"SELECT nome FROM user WHERE matricula = '{self.userg}' "
            ).fetchone()
            senha = self.c.execute(
                f"SELECT senha FROM user WHERE senha = '{self.senhag}' "
            ).fetchone()
            if senha != None:
                senha = senha[0]

            if self.senhag == senha:
                user = user[0]
                user = user.split()
                user = user[0]
                self.lblUser["text"] = f'👤 {user}'
                self.msg(1, "Login realizado com sucesso!")
                win2.destroy()

            else:
                self.msg(2, 'Senha ou Matricula Incorreta!')
            self.nome = user

        Label(win2,
              text="🧑 Matricula:",
              bg=self.bg,
              fg="white",
              font='arial 12 bold').pack(pady=5)
        self.user = Entry(win2, font='arial 12 bold', justify='center')
        self.user.pack(anchor="center")
        Label(win2,
              text="🔑 Senha:",
              bg=self.bg,
              fg="white",
              font='arial 12 bold').pack(pady=5)
        self.senha = Entry(win2,
                           font='arial 12 bold',
                           show='*',
                           justify='center')
        self.senha.pack(anchor="center")
        Button(win2, text="Login", font="arial 10", command=cons).pack(pady=10)
        win2.bind('<Return>', cons)
        win2.mainloop()

    def settings(self):
        if self.nome != 'Undefined':
            pass
        else:
            self.msg(3, "Realize Login Primeiro!!")

    def addUser(self):
        c2 = connect('vj.db')
        cv = c2.cursor()

        def add(event):
            nome = nomeCad.get().upper()
            mat = matCad.get()
            #try:
            cv.execute(
                f'INSERT INTO opDB(matricula, nome) VALUES ("{mat}","{nome}")')
            c2.commit()
            nome = nome.split()
            nome = nome[0]
            self.msg(1, f'{nome} adicionado com sucesso')
            winCad.destroy()
            #except:
            #self.msg(2, 'Erro!')

        if self.nome != 'Undefined':
            winCad = Tk()
            winCad.iconbitmap("img/icon.ico")
            winCad.geometry("250x400")
            winCad['bg'] = self.bg
            winCad.resizable(width=False, height=False)
            winCad.title("CADASTRO")

            Label(winCad, text='Nome', bg=self.bg, fg='white',
                  font='Arial 12').pack()
            nomeCad = Entry(winCad)
            nomeCad.pack()
            Label(winCad,
                  text='Matricula',
                  bg=self.bg,
                  fg='white',
                  font='Arial 12').pack()
            matCad = Entry(winCad)
            matCad.pack()
            matCad.bind('<Return>', add)
            Button(winCad, text='Cadastrar', command=partial(add, 1)).pack()

            winCad.mainloop()
        else:
            self.msg(3, "Realize Login Primeiro!!")

    def logout(self):
        if self.nome != 'Undefined':
            self.lblUser['text'] = 'Undefined'
            self.nome = 'Undefined'
            self.msg(1, 'Lougout realizado com sucesso!!')
        else:
            self.msg(1, 'Realize Login !!!')

    def caulcarGrafico(self):
        dia = int(st("%d"))
        self.lhr = st('%H:%M')
        pln = lw('metas.xlsx')
        pln2 = lw('horaxhora.xlsx')
        ws = pln.active
        ws2 = pln2.active
        vlr = []

        dcos = []
        for i in range(1, 6000):
            diaP = ws[f'A{i}'].value
            loja = ws[f'B{i}'].value
            dco = ws[f'C{i}'].value
            valor = ws[f'D{i}'].value
            if diaP != None:
                if loja == 'L062':
                    if diaP == dia:
                        dcos.append(dco)
                        vlr.append(valor)

        dep = []
        for d in dcos:
            dco = d[:4]
            dep.append(dco)

        plt.style.use('ggplot')
        plt.figure(figsize=(5, 2), facecolor='#333')
        plt.title('Metas Financeiras 062',
                  fontsize=16,
                  fontweight='bold',
                  color='white')
        meta = plt.subplot()
        meta.bar(dep, vlr)
        meta.set_facecolor('#333')
        plt.savefig('img\\meta.png')
        plt.close()

        y_pos = []
        deps = [
            'Beleza', 'Calça', 'Tecno', 'Fem', 'Inf', 'Masc', 'Moda C',
            'Óculos', 'Basket', 'Relóg'
        ]

        for i in range(1, 11):
            y_pos.append(i)

        realizado = []
        for i in range(7, 17):
            valor = ws2[f'E{i}'].value
            realizado.append(valor)

        plt.style.use('ggplot')
        plt.figure(figsize=(5, 2), facecolor='#333')
        plt.title(f'Realizado {self.lhr}',
                  fontsize=16,
                  fontweight='bold',
                  color='white')

        rel = plt.subplot()
        rel.barh(y_pos, realizado)
        rel.set_yticks(y_pos, deps)

        rel.set_facecolor('#333')
        plt.savefig('img\\real.png')
        plt.close()


App()
